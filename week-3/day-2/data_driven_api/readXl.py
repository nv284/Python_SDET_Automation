import requests
from openpyxl import load_workbook

file_path = r"week-3\\day-2\\data_driven_api\\users.xlsx"
wb = load_workbook(file_path)
sheet = wb.active
API_URL = "https://reqres.in/api/users"

print(" Starting API Tests using openpyxl...\n")

for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        break

username, job, expected_status = row
payload = {"name": username, "job": job}

try:
        response = requests.post(API_URL, json=payload)

        if response.status_code == int(expected_status):
            print(f" PASS | User: {username} | Status: {response.status_code}")
        else:
            print(
                f" FAIL | User: {username} | Expected {expected_status}, Got {response.status_code}"
            )
except Exception as e:
        print(f" ERROR executing request for {username}: {e}")

print("\n🏁 All test rows executed.")