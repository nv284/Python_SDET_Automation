import requests
from openpyxl import load_workbook
import io
from zipfile import ZipFile

# Load Excel file
file_path = r"week-3\\day-2\\data_driven_api\\users.xlsx"
workbook = load_workbook(file_path)
#workbook = load_workbook("users.xlsx")

sheet = workbook.active

# API URL
url = "https://jsonplaceholder.typicode.com/users"

# Read rows from Excel
for row in sheet.iter_rows(min_row=2, values_only=True):

    name, username, email = row

    print("\nExecuting API for:", name)

    payload = {
        "name": name,
        "username": username,
        "email": email
    }

    # Send POST request
    response = requests.post(url, json=payload)

    data = response.json()

    # Print response
    print("Status Code:", response.status_code)
    print("Response:", data)

    # Validations
    assert response.status_code == 201
    assert data["name"] == name

    print("Validation Passed")